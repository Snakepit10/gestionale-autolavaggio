from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Avg, Q
from django.utils import timezone
from datetime import datetime, timedelta, date
import json
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import csv

from .models import ReportPersonalizzato, EsecuzioneReport, Dashboard, KPI, StoricoCambiamenti
from apps.ordini.models import Ordine, ItemOrdine, Pagamento
from apps.clienti.models import Cliente
from apps.core.models import ServizioProdotto, Categoria, Postazione
from apps.abbonamenti.models import Abbonamento


class DashboardPrincipaleView(LoginRequiredMixin, DetailView):
    """Dashboard principale con KPI e statistiche"""
    template_name = 'reportistica/dashboard_principale.html'
    
    def get_object(self):
        # Ritorna un oggetto dummy per il template
        return type('obj', (object,), {'id': 1})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Periodo di default: ultimo mese
        fine = timezone.now().date()
        inizio = fine - timedelta(days=30)
        
        # KPI principali
        context['kpi_data'] = self.get_kpi_data(inizio, fine)
        
        # Grafici
        context['grafici_data'] = self.get_grafici_data(inizio, fine)
        
        # Ordini recenti
        context['ordini_recenti'] = Ordine.objects.select_related('cliente').order_by('-data_creazione')[:10]
        
        # Statistiche postazioni
        context['statistiche_postazioni'] = self.get_statistiche_postazioni(inizio, fine)
        
        return context
    
    def get_kpi_data(self, inizio, fine):
        """Calcola i KPI principali"""
        
        # Fatturato periodo
        fatturato_periodo = Pagamento.objects.filter(
            ordine__data_creazione__date__range=[inizio, fine],
            stato='completato'
        ).aggregate(totale=Sum('importo'))['totale'] or 0
        
        # Fatturato periodo precedente (per confronto)
        giorni_periodo = (fine - inizio).days
        inizio_precedente = inizio - timedelta(days=giorni_periodo)
        fine_precedente = inizio - timedelta(days=1)
        
        fatturato_precedente = Pagamento.objects.filter(
            ordine__data_creazione__date__range=[inizio_precedente, fine_precedente],
            stato='completato'
        ).aggregate(totale=Sum('importo'))['totale'] or 0
        
        # Calcola variazione percentuale
        if fatturato_precedente > 0:
            variazione_fatturato = ((fatturato_periodo - fatturato_precedente) / fatturato_precedente) * 100
        else:
            variazione_fatturato = 100 if fatturato_periodo > 0 else 0
        
        # Numero ordini
        ordini_periodo = Ordine.objects.filter(
            data_creazione__date__range=[inizio, fine]
        ).count()
        
        ordini_precedente = Ordine.objects.filter(
            data_creazione__date__range=[inizio_precedente, fine_precedente]
        ).count()
        
        if ordini_precedente > 0:
            variazione_ordini = ((ordini_periodo - ordini_precedente) / ordini_precedente) * 100
        else:
            variazione_ordini = 100 if ordini_periodo > 0 else 0
        
        # Nuovi clienti
        nuovi_clienti = Cliente.objects.filter(
            data_registrazione__date__range=[inizio, fine]
        ).count()
        
        # Scontrino medio
        if ordini_periodo > 0:
            scontrino_medio = fatturato_periodo / ordini_periodo
        else:
            scontrino_medio = 0
        
        return {
            'fatturato': {
                'valore': fatturato_periodo,
                'variazione': variazione_fatturato,
                'trend': 'up' if variazione_fatturato > 0 else 'down' if variazione_fatturato < 0 else 'stable'
            },
            'ordini': {
                'valore': ordini_periodo,
                'variazione': variazione_ordini,
                'trend': 'up' if variazione_ordini > 0 else 'down' if variazione_ordini < 0 else 'stable'
            },
            'nuovi_clienti': {
                'valore': nuovi_clienti,
                'variazione': 0,  # Placeholder
                'trend': 'stable'
            },
            'scontrino_medio': {
                'valore': scontrino_medio,
                'variazione': 0,  # Placeholder
                'trend': 'stable'
            }
        }
    
    def get_grafici_data(self, inizio, fine):
        """Prepara dati per i grafici"""
        
        # Fatturato giornaliero
        fatturato_giornaliero = []
        current_date = inizio
        while current_date <= fine:
            fatturato_giorno = Pagamento.objects.filter(
                ordine__data_creazione__date=current_date,
                stato='completato'
            ).aggregate(totale=Sum('importo'))['totale'] or 0
            
            fatturato_giornaliero.append({
                'data': current_date.isoformat(),
                'fatturato': float(fatturato_giorno)
            })
            current_date += timedelta(days=1)
        
        # Servizi più richiesti
        servizi_top = ItemOrdine.objects.filter(
            ordine__data_creazione__date__range=[inizio, fine]
        ).values('servizio__titolo').annotate(
            count=Count('id'),
            fatturato=Sum('subtotale')
        ).order_by('-count')[:10]
        
        # Distribuzione per categorie
        categorie_dist = ItemOrdine.objects.filter(
            ordine__data_creazione__date__range=[inizio, fine]
        ).values('servizio__categoria__nome').annotate(
            fatturato=Sum('subtotale')
        ).order_by('-fatturato')
        
        return {
            'fatturato_giornaliero': fatturato_giornaliero,
            'servizi_top': list(servizi_top),
            'categorie_distribuzione': list(categorie_dist)
        }
    
    def get_statistiche_postazioni(self, inizio, fine):
        """Statistiche utilizzo postazioni"""
        
        postazioni = Postazione.objects.filter(attiva=True)
        statistiche = []
        
        for postazione in postazioni:
            ordini_postazione = Ordine.objects.filter(
                postazione=postazione,
                data_creazione__date__range=[inizio, fine]
            )
            
            fatturato = ordini_postazione.aggregate(
                totale=Sum('pagamenti__importo')
            )['totale'] or 0
            
            statistiche.append({
                'postazione': postazione,
                'ordini': ordini_postazione.count(),
                'fatturato': fatturato,
                'utilizzo_percentuale': 85  # Placeholder
            })
        
        return statistiche


class ReportListView(LoginRequiredMixin, ListView):
    """Lista dei report disponibili"""
    model = ReportPersonalizzato
    template_name = 'reportistica/report_list.html'
    context_object_name = 'reports'
    paginate_by = 20
    
    def get_queryset(self):
        return ReportPersonalizzato.objects.filter(
            Q(creato_da=self.request.user) | Q(creato_da__is_staff=True)
        ).order_by('-data_creazione')


class ReportDetailView(LoginRequiredMixin, DetailView):
    """Dettaglio e configurazione report"""
    model = ReportPersonalizzato
    template_name = 'reportistica/report_detail.html'
    context_object_name = 'report'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Ultime esecuzioni
        context['esecuzioni'] = self.object.esecuzioni.order_by('-data_esecuzione')[:10]
        
        return context


@login_required
def genera_report(request, report_id):
    """Genera un report specifico"""
    report = get_object_or_404(ReportPersonalizzato, id=report_id)
    
    # Parametri dal form
    data_inizio = request.POST.get('data_inizio')
    data_fine = request.POST.get('data_fine')
    formato = request.POST.get('formato', report.formato_default)
    
    if not data_inizio or not data_fine:
        return JsonResponse({'error': 'Date mancanti'}, status=400)
    
    try:
        data_inizio = datetime.strptime(data_inizio, '%Y-%m-%d').date()
        data_fine = datetime.strptime(data_fine, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Formato date non valido'}, status=400)
    
    # Crea esecuzione
    esecuzione = EsecuzioneReport.objects.create(
        report=report,
        eseguito_da=request.user,
        parametri_esecuzione={
            'data_inizio': data_inizio.isoformat(),
            'data_fine': data_fine.isoformat(),
            'formato': formato
        }
    )
    
    try:
        # Genera il report in base al tipo
        if report.tipo == 'vendite':
            response = genera_report_vendite(data_inizio, data_fine, formato)
        elif report.tipo == 'clienti':
            response = genera_report_clienti(data_inizio, data_fine, formato)
        elif report.tipo == 'servizi':
            response = genera_report_servizi(data_inizio, data_fine, formato)
        else:
            raise ValueError(f"Tipo report non supportato: {report.tipo}")
        
        # Aggiorna esecuzione
        esecuzione.stato = 'completato'
        esecuzione.data_completamento = timezone.now()
        esecuzione.save()
        
        return response
        
    except Exception as e:
        esecuzione.stato = 'errore'
        esecuzione.messaggio_errore = str(e)
        esecuzione.save()
        
        return JsonResponse({'error': str(e)}, status=500)


def genera_report_vendite(data_inizio, data_fine, formato='pdf'):
    """Genera report vendite"""
    
    # Query dati
    ordini = Ordine.objects.filter(
        data_creazione__date__range=[data_inizio, data_fine]
    ).select_related('cliente', 'postazione')
    
    # Calcoli aggregati
    totale_fatturato = ordini.aggregate(
        totale=Sum('pagamenti__importo')
    )['totale'] or 0
    
    numero_ordini = ordini.count()
    scontrino_medio = totale_fatturato / numero_ordini if numero_ordini > 0 else 0
    
    if formato == 'pdf':
        return genera_pdf_vendite(ordini, totale_fatturato, numero_ordini, scontrino_medio, data_inizio, data_fine)
    elif formato == 'excel':
        return genera_excel_vendite(ordini, totale_fatturato, numero_ordini, scontrino_medio, data_inizio, data_fine)
    elif formato == 'csv':
        return genera_csv_vendite(ordini, data_inizio, data_fine)


def genera_pdf_vendite(ordini, totale_fatturato, numero_ordini, scontrino_medio, data_inizio, data_fine):
    """Genera PDF report vendite"""
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Titolo
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.darkblue,
        alignment=1  # Centrato
    )
    
    story.append(Paragraph("Report Vendite", title_style))
    story.append(Paragraph(f"Periodo: {data_inizio.strftime('%d/%m/%Y')} - {data_fine.strftime('%d/%m/%Y')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Riassunto
    summary_data = [
        ['Indicatore', 'Valore'],
        ['Totale Fatturato', f'€ {totale_fatturato:.2f}'],
        ['Numero Ordini', str(numero_ordini)],
        ['Scontrino Medio', f'€ {scontrino_medio:.2f}'],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 30))
    
    # Dettaglio ordini
    story.append(Paragraph("Dettaglio Ordini", styles['Heading2']))
    story.append(Spacer(1, 10))
    
    # Tabella ordini (prime 50)
    ordini_data = [['Data', 'N. Ordine', 'Cliente', 'Postazione', 'Totale']]
    
    for ordine in ordini[:50]:  # Limita per non appesantire il PDF
        cliente_nome = ordine.cliente.nome_completo if ordine.cliente else 'Anonimo'
        postazione_nome = ordine.postazione.nome if ordine.postazione else '-'
        totale_ordine = ordine.totale or 0
        
        ordini_data.append([
            ordine.data_creazione.strftime('%d/%m/%Y'),
            ordine.numero_ordine,
            cliente_nome[:30],  # Tronca se troppo lungo
            postazione_nome,
            f'€ {totale_ordine:.2f}'
        ])
    
    ordini_table = Table(ordini_data, colWidths=[1*inch, 1*inch, 2*inch, 1*inch, 1*inch])
    ordini_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(ordini_table)
    
    # Genera PDF
    doc.build(story)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="report_vendite_{data_inizio}_{data_fine}.pdf"'
    
    return response


def genera_excel_vendite(ordini, totale_fatturato, numero_ordini, scontrino_medio, data_inizio, data_fine):
    """Genera Excel report vendite"""
    
    # Crea workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Report Vendite"
    
    # Stili
    header_font = Font(bold=True, size=14)
    normal_font = Font(size=11)
    
    # Intestazioni
    ws['A1'] = "REPORT VENDITE"
    ws['A1'].font = header_font
    ws['A2'] = f"Periodo: {data_inizio.strftime('%d/%m/%Y')} - {data_fine.strftime('%d/%m/%Y')}"
    
    # Riassunto
    ws['A4'] = "RIASSUNTO"
    ws['A4'].font = header_font
    ws['A5'] = "Totale Fatturato:"
    ws['B5'] = totale_fatturato
    ws['A6'] = "Numero Ordini:"
    ws['B6'] = numero_ordini
    ws['A7'] = "Scontrino Medio:"
    ws['B7'] = scontrino_medio
    
    # Intestazioni dettaglio
    row = 10
    headers = ['Data', 'N. Ordine', 'Cliente', 'Postazione', 'Totale']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
    
    # Dati ordini
    for ordine in ordini:
        row += 1
        cliente_nome = ordine.cliente.nome_completo if ordine.cliente else 'Anonimo'
        postazione_nome = ordine.postazione.nome if ordine.postazione else '-'
        totale_ordine = ordine.totale or 0
        
        ws.cell(row=row, column=1, value=ordine.data_creazione.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=ordine.numero_ordine)
        ws.cell(row=row, column=3, value=cliente_nome)
        ws.cell(row=row, column=4, value=postazione_nome)
        ws.cell(row=row, column=5, value=totale_ordine)
    
    # Salva in buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="report_vendite_{data_inizio}_{data_fine}.xlsx"'
    
    return response


def genera_csv_vendite(ordini, data_inizio, data_fine):
    """Genera CSV report vendite"""
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="report_vendite_{data_inizio}_{data_fine}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Data', 'N. Ordine', 'Cliente', 'Postazione', 'Totale'])
    
    for ordine in ordini:
        cliente_nome = ordine.cliente.nome_completo if ordine.cliente else 'Anonimo'
        postazione_nome = ordine.postazione.nome if ordine.postazione else '-'
        totale_ordine = ordine.totale or 0
        
        writer.writerow([
            ordine.data_creazione.strftime('%d/%m/%Y'),
            ordine.numero_ordine,
            cliente_nome,
            postazione_nome,
            f'{totale_ordine:.2f}'
        ])
    
    return response


def genera_report_clienti(data_inizio, data_fine, formato='pdf'):
    """Genera report clienti - placeholder"""
    if formato == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="report_clienti_{data_inizio}_{data_fine}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Nome', 'Email', 'Data Registrazione', 'Ordini Totali'])
        
        clienti = Cliente.objects.filter(
            data_registrazione__date__range=[data_inizio, data_fine]
        )
        
        for cliente in clienti:
            writer.writerow([
                cliente.nome_completo,
                cliente.email,
                cliente.data_registrazione.strftime('%d/%m/%Y'),
                cliente.ordini.count()
            ])
        
        return response
    
    # Per ora solo CSV implementato
    return JsonResponse({'error': 'Formato non supportato per report clienti'}, status=400)


def genera_report_servizi(data_inizio, data_fine, formato='pdf'):
    """Genera report servizi - placeholder"""
    if formato == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="report_servizi_{data_inizio}_{data_fine}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Servizio', 'Categoria', 'Volte Venduto', 'Fatturato Totale'])
        
        # Query servizi più venduti
        servizi_stats = ItemOrdine.objects.filter(
            ordine__data_creazione__date__range=[data_inizio, data_fine]
        ).values(
            'servizio__titolo',
            'servizio__categoria__nome'
        ).annotate(
            volte_venduto=Count('id'),
            fatturato_totale=Sum('subtotale')
        ).order_by('-volte_venduto')
        
        for stat in servizi_stats:
            writer.writerow([
                stat['servizio__titolo'],
                stat['servizio__categoria__nome'],
                stat['volte_venduto'],
                f"{stat['fatturato_totale']:.2f}"
            ])
        
        return response
    
    # Per ora solo CSV implementato
    return JsonResponse({'error': 'Formato non supportato per report servizi'}, status=400)


@login_required
def api_kpi_data(request):
    """API per dati KPI in tempo reale"""
    
    # Oggi
    oggi = timezone.now().date()
    
    # Fatturato oggi
    fatturato_oggi = Pagamento.objects.filter(
        ordine__data_creazione__date=oggi,
        stato='completato'
    ).aggregate(totale=Sum('importo'))['totale'] or 0
    
    # Ordini oggi
    ordini_oggi = Ordine.objects.filter(
        data_creazione__date=oggi
    ).count()
    
    # Clienti attivi (con ordini negli ultimi 30 giorni)
    clienti_attivi = Cliente.objects.filter(
        ordini__data_creazione__date__gte=oggi - timedelta(days=30)
    ).distinct().count()
    
    return JsonResponse({
        'fatturato_oggi': float(fatturato_oggi),
        'ordini_oggi': ordini_oggi,
        'clienti_attivi': clienti_attivi,
        'timestamp': timezone.now().isoformat()
    })


@login_required
def esporta_dati_completi(request):
    """Esporta tutti i dati in formato JSON per backup"""
    
    if not request.user.is_staff:
        return JsonResponse({'error': 'Permessi insufficienti'}, status=403)
    
    from django.core import serializers
    from django.apps import apps
    
    data = {}
    
    # Lista modelli da esportare
    modelli_export = [
        'core.Categoria',
        'core.ServizioProdotto', 
        'clienti.Cliente',
        'ordini.Ordine',
        'ordini.ItemOrdine',
        'abbonamenti.Abbonamento',
    ]
    
    for modello_path in modelli_export:
        app_label, model_name = modello_path.split('.')
        model = apps.get_model(app_label, model_name)
        
        serialized = serializers.serialize('json', model.objects.all())
        data[modello_path] = json.loads(serialized)
    
    response = JsonResponse(data)
    response['Content-Disposition'] = f'attachment; filename="backup_dati_{oggi.isoformat()}.json"'
    
    return response